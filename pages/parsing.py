from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from os.path import join, abspath

def working_with_table(file_name, tabs_number, string_tabl, column_tabl):
    # excel файл должен быть в этой же директории
    add_files = abspath(join('.', file_name))
    # Настраиваем опции документа
    file_option = load_workbook(filename=add_files, data_only=True, read_only=True)

    # Получаем список с вкладками документа - sheetnames
    tabs = list(file_option.sheetnames)
    # Определяем нужную вкладку, икорректируем номер указания в функции. Зачем нам начинать с нуля? :)
    target_tabs = tabs[tabs_number - 1]

    # Обращаемся к нужной вкладке
    employees_sheet = file_option[f'{target_tabs}']

    list_value = list()
    while employees_sheet.cell(string_tabl, column_tabl).value is not None:
        list_value.append(employees_sheet.cell(string_tabl, column_tabl).value)
        string_tabl += 1
    return list_value

'''Списки входных данных'''
class ListParsingTabs:
    car_type = working_with_table('ExLab_task.xlsx', 2, 3, 2)       # тип автомобиля
    body_type = working_with_table('ExLab_task.xlsx', 2, 3, 3)      # тип кузова
    car_brand = working_with_table('ExLab_task.xlsx', 2, 3, 4)      # марка
    fuel_for_car = working_with_table('ExLab_task.xlsx', 2, 3, 5)   # тип топлива
    transmission = working_with_table('ExLab_task.xlsx', 2, 3, 6)   # тип коробки передач
    car_color = working_with_table('ExLab_task.xlsx', 2, 3, 7)      # цвет машины

print(ListParsingTabs.car_type)


